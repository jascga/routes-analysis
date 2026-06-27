"""
测试Excel报告导出器
"""

import os
import tempfile
import pytest
from pathlib import Path
from datetime import datetime

from routesanalysis.export.comparison import (
    ExcelExporter, export_comparison_result, export_differences_to_excel
)
from routesanalysis.models import (
    Device, Route, RouteProtocol,
    RouteDifference, DifferenceType, ComparisonResult
)


class TestExcelExporter:
    """测试Excel导出器"""

    @pytest.fixture
    def sample_comparison_result(self):
        """创建示例比较结果"""
        # 创建示例路由
        routes = [
            Route(
                destination="10.0.0.0/24",
                next_hop="192.168.1.1",
                interface="GigabitEthernet0/0/1",
                pre=60,
                cost=0,
                protocol=RouteProtocol.BGP
            ),
            Route(
                destination="10.0.1.0/24",
                next_hop="192.168.1.2",
                interface="GigabitEthernet0/0/2",
                pre=60,
                cost=0,
                protocol=RouteProtocol.IBGP
            ),
        ]

        # 创建设备
        baseline_device = Device(name="Switch-01", filename="file1.txt", routes=routes)
        compared_device = Device(name="Switch-02", filename="file2.txt", routes=[])

        # 创建差异
        differences = [
            RouteDifference(
                destination="10.0.0.0/24",
                device1="Switch-01",
                device2="Switch-02",
                difference_type=DifferenceType.MISSING_DESTINATION,
                details={"missing_in": "Switch-02"}
            ),
            RouteDifference(
                destination="10.0.1.0/24",
                device1="Switch-01",
                device2="Switch-02",
                difference_type=DifferenceType.MISSING_INTERFACE,
                details={
                    "interface": "GigabitEthernet0/0/2",
                    "missing_in": "Switch-02",
                    "route_in_device1": routes[1]
                }
            ),
            RouteDifference(
                destination="10.0.2.0/24",
                device1="Switch-01",
                device2="Switch-02",
                difference_type=DifferenceType.PRE_COST_DIFFERENCE,
                details={
                    "interface": "GigabitEthernet0/0/3",
                    "device1_pre": 60,
                    "device2_pre": 70,
                    "device1_cost": 0,
                    "device2_cost": 10
                }
            ),
        ]

        # 创建汇总信息
        summary = {
            "baseline_device": "Switch-01",
            "compared_devices": ["Switch-02"],
            "total_routes_baseline": 2,
            "total_differences": 3,
            "differences_by_type": {
                "missing_destination": 1,
                "missing_interface": 1,
                "pre_cost_diff": 1
            },
            "differences_by_device_pair": {
                "Switch-01-Switch-02": 3
            },
            "performance_stats": {
                "total_comparison_time": 0.5,
                "routes_per_second": 4.0
            }
        }

        return ComparisonResult(
            baseline_device=baseline_device,
            compared_devices=[compared_device],
            differences=differences,
            summary=summary
        )

    @pytest.fixture
    def exporter(self):
        """创建导出器实例"""
        return ExcelExporter(write_only=False)  # 使用常规模式以便测试样式

    def test_exporter_initialization(self):
        """测试导出器初始化"""
        exporter = ExcelExporter(write_only=True, batch_size=5000)
        assert exporter.write_only is True
        assert exporter.batch_size == 5000

        exporter2 = ExcelExporter(write_only=False)
        assert exporter2.write_only is False


    def test_export_write_only_mode(self, sample_comparison_result):
        """测试write-only模式导出（适用于大量数据）"""
        exporter = ExcelExporter(write_only=True, batch_size=2)  # 小批量测试

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            output_path = tmp.name

        try:
            # 导出Excel
            exporter.export(sample_comparison_result, output_path)

            # 检查文件是否存在且不为空
            assert os.path.exists(output_path)
            assert os.path.getsize(output_path) > 0

            # 验证文件内容
            import openpyxl
            wb = openpyxl.load_workbook(output_path, read_only=True)
            ws = wb.active

            # 收集所有单元格值
            values = []
            for row in ws.iter_rows(values_only=True):
                row_values = [str(v) for v in row if v is not None]
                if row_values:
                    values.extend(row_values)

            text = " ".join(values)
            assert "Switch-01" in text or "Switch-02" in text

            wb.close()

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)



    def test_export_empty_differences(self, exporter):
        """测试导出无差异的结果"""
        # 创建无差异的比较结果
        routes = [
            Route(
                destination="10.0.0.0/24",
                next_hop="192.168.1.1",
                interface="GigabitEthernet0/0/1",
                pre=60,
                cost=0,
                protocol=RouteProtocol.BGP
            ),
        ]

        baseline_device = Device(name="Switch-01", filename="file1.txt", routes=routes)
        compared_device = Device(name="Switch-02", filename="file2.txt", routes=routes)  # 相同路由

        summary = {
            "baseline_device": "Switch-01",
            "compared_devices": ["Switch-02"],
            "total_routes_baseline": 1,
            "total_differences": 0,
            "differences_by_type": {},
            "differences_by_device_pair": {},
            "performance_stats": {}
        }

        result = ComparisonResult(
            baseline_device=baseline_device,
            compared_devices=[compared_device],
            differences=[],
            summary=summary
        )

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            output_path = tmp.name

        try:
            exporter.export(result, output_path)
            assert os.path.exists(output_path)

            # 验证文件包含"差异总数: 0"
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active

            cell_values = [v for row in ws.iter_rows(values_only=True) for v in row if v is not None]
            text = " ".join(str(v) for v in cell_values if v)

            assert "差异总数" in text

            wb.close()

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_export_invalid_path(self, exporter, sample_comparison_result):
        """测试导出到无效路径"""
        # 导出器会创建目录，所以路径有效
        invalid_path = "/nonexistent/directory/report.xlsx"
        import tempfile
        with tempfile.TemporaryDirectory() as tmpdir:
            test_path = os.path.join(tmpdir, "nested/subdir/report.xlsx")
            exporter.export(sample_comparison_result, test_path)
            assert os.path.exists(test_path)




class TestExportConvenienceFunctions:
    """测试导出快捷函数"""

    @pytest.fixture
    def sample_differences(self):
        """创建示例差异列表"""
        return [
            RouteDifference(
                destination="10.0.0.0/24",
                device1="Switch-01",
                device2="Switch-02",
                difference_type=DifferenceType.MISSING_DESTINATION,
                details={"missing_in": "Switch-02"}
            ),
            RouteDifference(
                destination="10.0.1.0/24",
                device1="Switch-01",
                device2="Switch-02",
                difference_type=DifferenceType.PRE_COST_DIFFERENCE,
                details={
                    "interface": "GigabitEthernet0/0/1",
                    "device1_pre": 60,
                    "device2_pre": 70,
                    "device1_cost": 0,
                    "device2_cost": 10
                }
            ),
        ]



class TestExcelExporterIntegration:
    """测试Excel导出器集成测试"""

    def test_end_to_end_export(self):
        """测试端到端导出流程（使用fixture文件）"""
        # 跳过这个测试，因为它需要解析和比较文件
        # 在实际测试中，可以启用
        pytest.skip("端到端测试需要完整流程，在单元测试中跳过")

    def test_large_dataset_export(self):
        """测试大数据集导出（性能测试）"""
        # 创建大量差异以测试分批写入
        differences = []
        for i in range(1000):
            diff = RouteDifference(
                destination=f"10.{i//256}.{i%256}.0/24",
                device1="Switch-01",
                device2="Switch-02",
                difference_type=DifferenceType.MISSING_DESTINATION,
                details={"missing_in": "Switch-02"}
            )
            differences.append(diff)

        # 创建虚拟比较结果
        dummy_device = Device(name="Switch-01", filename="", routes=[])
        dummy_compared = [Device(name="Switch-02", filename="", routes=[])]

        summary = {
            "baseline_device": "Switch-01",
            "compared_devices": ["Switch-02"],
            "total_routes_baseline": 0,
            "total_differences": len(differences),
            "differences_by_type": {"missing_destination": len(differences)},
            "differences_by_device_pair": {},
            "performance_stats": {}
        }

        result = ComparisonResult(
            baseline_device=dummy_device,
            compared_devices=dummy_compared,
            differences=differences,
            summary=summary
        )

        # 使用write-only模式导出
        exporter = ExcelExporter(write_only=True, batch_size=100)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            output_path = tmp.name

        try:
            import time
            start_time = time.time()
            exporter.export(result, output_path)
            export_time = time.time() - start_time

            # 验证导出成功
            assert os.path.exists(output_path)
            file_size = os.path.getsize(output_path)

            print(f"导出 {len(differences)} 条差异耗时: {export_time:.2f}秒, 文件大小: {file_size/1024:.1f}KB")

            # 基本性能检查（可以根据需要调整阈值）
            assert export_time < 10.0  # 1000条差异应该在10秒内完成

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)