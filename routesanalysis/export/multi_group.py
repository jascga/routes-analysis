"""导出层：multi-group 场景的 Excel 报告"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from ..analyzer import MultiGroupAnalysisResult
from .base import (
    style_header, style_body_row, autosize,
    WARN_FILL, HIT_FILL,
    ensure_output_path,
)


def export_multi_group_result(result: MultiGroupAnalysisResult, output_path: str | Path) -> Path:
    """导出场景 1 分析结果到 Excel"""
    output_path = ensure_output_path(output_path)

    wb = Workbook()
    _write_summary_sheet(wb, result)
    _write_hits_sheet(wb, result)
    _write_all_routes_sheet(wb, result)
    _write_groups_sheet(wb, result)

    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 and wb["Sheet"].cell(1, 1).value is None:
        del wb["Sheet"]

    wb.save(output_path)
    return output_path


def _write_summary_sheet(wb: Workbook, result: MultiGroupAnalysisResult):
    ws = wb.create_sheet("汇总", 0)

    rows = [
        ("设备名", result.device.name),
        ("源文件", result.device.filename),
        ("BGP 路由总数", result.total_routes),
        ("目的网段数 (Destination)", result.total_destinations),
        ("命中路由数 (负载分担到 ≥ {} 组平行设备)".format(result.min_groups), result.hit_count),
        ("识别出的平行设备组数", len(result.group_members)),
        ("不规范对端设备名 (单独成组)", len(result.unparseable_peers)),
        ("是否包含接口描述", "是" if result.device.has_interface_descriptions() else "否（无对端映射，结果可能不准确）"),
        ("最小命中分组数 (min-groups)", result.min_groups),
    ]

    ws.cell(row=1, column=1, value="项目")
    ws.cell(row=1, column=2, value="值")
    style_header(ws, 1, 2)

    for i, (key, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=value)
        warn = (key.startswith("是否包含接口描述") and value != "是") or \
               (key.startswith("不规范对端") and isinstance(value, int) and value > 0)
        style_body_row(ws, i, 2, WARN_FILL if warn else None)

    autosize(ws, max_width=80)


def _write_hits_sheet(wb: Workbook, result: MultiGroupAnalysisResult):
    ws = wb.create_sheet("命中明细")

    headers = [
        "Destination", "路径数", "涉及组数", "涉及的平行设备组",
        "对端设备明细", "接口明细", "Pre", "Cost", "协议", "包含不规范对端",
    ]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    style_header(ws, 1, len(headers))

    for i, hit in enumerate(result.hits, start=2):
        ws.cell(row=i, column=1, value=hit.destination)
        ws.cell(row=i, column=2, value=hit.path_count)
        ws.cell(row=i, column=3, value=hit.group_count)
        ws.cell(row=i, column=4, value="\n".join(hit.group_keys))
        ws.cell(row=i, column=5, value="\n".join(f"{p.group_key} / {p.peer_device}" for p in hit.paths))
        ws.cell(row=i, column=6, value="\n".join(f"{p.peer_device} ← {p.interface}" for p in hit.paths))
        ws.cell(row=i, column=7, value=",".join(sorted({str(p.pre) for p in hit.paths})))
        ws.cell(row=i, column=8, value=",".join(sorted({str(p.cost) for p in hit.paths})))
        ws.cell(row=i, column=9, value=",".join(hit.protocols))
        ws.cell(row=i, column=10, value="是" if hit.has_unparseable_peer else "")

        fill = WARN_FILL if hit.has_unparseable_peer else HIT_FILL
        style_body_row(ws, i, len(headers), fill)

    ws.freeze_panes = "A2"
    autosize(ws, max_width=70)


def _write_all_routes_sheet(wb: Workbook, result: MultiGroupAnalysisResult):
    """所有 Destination 及其对应的分组信息"""
    ws = wb.create_sheet("所有路由")

    headers = [
        "Destination", "路径数", "涉及组数", "涉及的平行设备组",
        "是否命中", "对端设备明细", "接口明细", "Pre", "Cost", "协议",
    ]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    style_header(ws, 1, len(headers))

    hit_destinations = {h.destination for h in result.hits}

    row = 2
    for dest, paths in result.all_destinations.items():
        is_hit = dest in hit_destinations
        unique_groups = sorted({p.group_key for p in paths})

        ws.cell(row=row, column=1, value=dest)
        ws.cell(row=row, column=2, value=len(paths))
        ws.cell(row=row, column=3, value=len(unique_groups))
        ws.cell(row=row, column=4, value="\n".join(unique_groups))
        ws.cell(row=row, column=5, value="是" if is_hit else "")
        ws.cell(row=row, column=6, value="\n".join(f"{p.group_key} / {p.peer_device}" for p in paths))
        ws.cell(row=row, column=7, value="\n".join(f"{p.peer_device} ← {p.interface}" for p in paths))
        ws.cell(row=row, column=8, value=",".join(sorted({str(p.pre) for p in paths})))
        ws.cell(row=row, column=9, value=",".join(sorted({str(p.cost) for p in paths})))
        ws.cell(row=row, column=10, value=",".join(sorted({p.protocol for p in paths})))

        fill = HIT_FILL if is_hit else None
        style_body_row(ws, row, len(headers), fill)
        row += 1

    ws.freeze_panes = "A2"
    autosize(ws, max_width=70)


def _write_groups_sheet(wb: Workbook, result: MultiGroupAnalysisResult):
    ws = wb.create_sheet("平行设备组清单")

    headers = ["分组键", "成员数", "成员设备", "规范"]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    style_header(ws, 1, len(headers))

    sorted_groups = sorted(
        result.group_members.items(),
        key=lambda kv: (-len(kv[1]), kv[0]),
    )

    for i, (group_key, members) in enumerate(sorted_groups, start=2):
        ws.cell(row=i, column=1, value=group_key)
        ws.cell(row=i, column=2, value=len(members))
        ws.cell(row=i, column=3, value="\n".join(sorted(members)))
        is_unparseable_only = all(m in result.unparseable_peers for m in members)
        ws.cell(row=i, column=4, value="否" if is_unparseable_only else "是")
        style_body_row(ws, i, len(headers), WARN_FILL if is_unparseable_only else None)

    ws.freeze_panes = "A2"
    autosize(ws, max_width=60)
