"""导出层：comparison 场景的 Excel 差异报告"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from ..models import ComparisonResult
from .base import (
    style_header, style_body_row, autosize,
    WARN_FILL,
    ensure_output_path,
)


def export_comparison_result(result: ComparisonResult, output_path: str | Path) -> Path:
    """导出场景 2 比较结果到 Excel（快捷函数）"""
    return ExcelExporter().export(result, output_path)


class ExcelExporter:
    """Excel 差异报告导出器（兼容旧版 API）"""

    def __init__(self, write_only: bool = False, batch_size: int = 10000):
        self.write_only = write_only
        self.batch_size = batch_size

    def export(self, result: ComparisonResult, output_path: str | Path) -> Path:
        output_path = ensure_output_path(output_path)
        wb = Workbook()
        _write_summary_sheet(wb, result)
        _write_details_sheet(wb, result)

        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 and wb["Sheet"].cell(1, 1).value is None:
            del wb["Sheet"]

        wb.save(output_path)
        return output_path


_DIFF_TYPE_NAME = {
    "missing_destination": "缺少Destination",
    "missing_interface": "接口/对端缺失",
    "interface_mismatch": "接口不同",
    "pre_cost_diff": "Pre/Cost差异",
}

_DIFF_TYPE_FILL = {
    "missing_destination": "FFC7CE",   # 红
    "missing_interface": "FFE699",     # 黄
    "interface_mismatch": "F4B183",    # 橙
    "pre_cost_diff": "C6EFCE",         # 绿
}


def _write_summary_sheet(wb: Workbook, result: ComparisonResult):
    ws = wb.create_sheet("汇总", 0)

    stats = result.get_statistics()
    by_type = stats.get("by_type", {})

    rows = [
        ("基准设备", result.baseline_device.name),
        ("基准设备路由数", len(result.baseline_device.routes)),
        ("比较设备", ", ".join(d.name for d in result.compared_devices)),
        ("差异总数", len(result.differences)),
    ]
    for k, v in sorted(by_type.items()):
        rows.append((_DIFF_TYPE_NAME.get(k, k), v))

    ws.cell(row=1, column=1, value="项目")
    ws.cell(row=1, column=2, value="值")
    style_header(ws, 1, 2)

    for i, (key, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=value)
        style_body_row(ws, i, 2)

    autosize(ws, max_width=80)


def _write_details_sheet(wb: Workbook, result: ComparisonResult):
    ws = wb.create_sheet("差异明细")

    headers = [
        "Destination", "差异类型", "设备1", "设备2", "详情",
    ]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    style_header(ws, 1, len(headers))

    for i, diff in enumerate(result.differences, start=2):
        diff_type = diff.difference_type.value
        details = diff.details

        ws.cell(row=i, column=1, value=diff.destination)
        ws.cell(row=i, column=2, value=_DIFF_TYPE_NAME.get(diff_type, diff_type))
        ws.cell(row=i, column=3, value=diff.device1)
        ws.cell(row=i, column=4, value=diff.device2)

        # 详情文本
        if diff_type == "missing_destination":
            detail_text = f"在 {details.get('missing_in', '?')} 中缺失"
        elif diff_type == "missing_interface":
            label = "对端设备" if details.get("compared_by") == "peer_device" else "接口"
            detail_text = f"{label} {details.get('interface', '?')} 在 {details.get('missing_in', '?')} 中缺失"
        elif diff_type == "interface_mismatch":
            detail_text = (f"设备1: [{', '.join(details.get('device1_interfaces', []))}] "
                          f"vs 设备2: [{', '.join(details.get('device2_interfaces', []))}]")
        else:  # pre_cost_diff
            detail_text = (f"接口 {details.get('interface', '?')}: "
                          f"Pre({details.get('device1_pre', '?')} vs {details.get('device2_pre', '?')}), "
                          f"Cost({details.get('device1_cost', '?')} vs {details.get('device2_cost', '?')})")
        ws.cell(row=i, column=5, value=detail_text)

        fill_color = _DIFF_TYPE_FILL.get(diff_type)
        from openpyxl.styles import PatternFill
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid") if fill_color else None
        style_body_row(ws, i, len(headers), fill)

    ws.freeze_panes = "A2"
    autosize(ws, max_width=80)


def export_differences_to_excel(differences, output_path,
                               baseline_name="设备1",
                               compared_name="设备2"):
    """将差异列表导出到Excel（快捷函数，兼容旧版API）"""
    from ..models import Device, ComparisonResult
    from collections import Counter

    dummy_device = Device(name=baseline_name, filename="", routes=[])
    dummy_compared = [Device(name=compared_name, filename="", routes=[])]

    diff_counter = Counter()
    for diff in differences:
        diff_counter[diff.difference_type] += 1

    summary = {
        "baseline_device": baseline_name,
        "compared_devices": [compared_name],
        "total_routes_baseline": 0,
        "total_differences": len(differences),
        "differences_by_type": {k.value: v for k, v in diff_counter.items()},
        "differences_by_device_pair": {},
        "performance_stats": {},
    }

    result = ComparisonResult(
        baseline_device=dummy_device,
        compared_devices=dummy_compared,
        differences=differences,
        summary=summary,
    )
    return export_comparison_result(result, output_path)
