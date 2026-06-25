"""导出层公共样式和工具函数"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 公共样式常量
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

WARN_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
HIT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

BODY_FONT = Font(name="微软雅黑", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# 公共工具函数
# ---------------------------------------------------------------------------

def style_header(ws, row: int, last_col: int):
    """给表头行设置样式"""
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_body_row(ws, row: int, last_col: int, fill: Optional[PatternFill] = None):
    """给数据行设置样式"""
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = BORDER
        if fill is not None:
            cell.fill = fill


def autosize(ws, max_width: int = 60):
    """自动调整列宽"""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            for line in str(cell.value).splitlines():
                length = max(length, _display_width(line))
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), max_width)


def _display_width(s: str) -> int:
    """简单估算字符宽度：中文按 2，其他按 1"""
    return sum(2 if ord(c) > 127 else 1 for c in s)


def ensure_output_path(output_path: str | Path) -> Path:
    """确保输出目录存在"""
    p = Path(output_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    return p
