"""
Excel 导出器 - 场景 1：多组平行设备负载分担分析

输出 3 个 Sheet：
- 汇总：设备信息、分析参数、统计指标
- 命中明细：每条命中的路由，含路径、对端、分组等
- 平行设备组清单：分组键 -> 该组下出现过的对端设备列表
"""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .analyzer import MultiGroupAnalysisResult


# ---------------------------------------------------------------------------
# 样式
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

_WARN_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
_HIT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

_BODY_FONT = Font(name="微软雅黑", size=10)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(ws, row: int, last_col: int):
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER


def _style_body_row(ws, row: int, last_col: int, fill: PatternFill | None = None):
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _BODY_FONT
        cell.alignment = _LEFT
        cell.border = _BORDER
        if fill is not None:
            cell.fill = fill


def _autosize(ws, max_width: int = 60):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            for line in str(cell.value).splitlines():
                length = max(length, _display_width(line))
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), max_width)


def _display_width(s: str) -> int:
    """简单估算字符宽度：中文按 2，其他按 1"""
    return sum(2 if ord(c) > 127 else 1 for c in s)


# ---------------------------------------------------------------------------
# 主导出函数
# ---------------------------------------------------------------------------

def export_multi_group_result(result: MultiGroupAnalysisResult, output_path: str | Path) -> Path:
    """导出场景 1 分析结果到 Excel"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _write_summary_sheet(wb, result)
    _write_hits_sheet(wb, result)
    _write_groups_sheet(wb, result)

    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 and wb["Sheet"].cell(1, 1).value is None:
        del wb["Sheet"]

    wb.save(output_path)
    return output_path


def _write_summary_sheet(wb: Workbook, result: MultiGroupAnalysisResult):
    ws = wb.create_sheet("汇总", 0)

    rows = [
        ("设备名", result.device.name),
        ("源文件", result.device.filename),
        ("BGP 路由总数", result.total_routes),
        ("目的网段数 (Destination)", result.total_destinations),
        ("命中路由数 (负载分担到 ≥ {} 组平行设备)".format(result.min_groups), result.hit_count),
        ("识别出的平行设备组数", len(result.group_members)),
        ("不规范对端设备名 (单独成组)", len(result.unparseable_peers)),
        ("是否包含接口描述", "是" if result.device.has_interface_descriptions() else "否（无对端映射，结果可能不准确）"),
        ("最小命中分组数 (min-groups)", result.min_groups),
    ]

    ws.cell(row=1, column=1, value="项目")
    ws.cell(row=1, column=2, value="值")
    _style_header(ws, 1, 2)

    for i, (key, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=value)
        warn = (key.startswith("是否包含接口描述") and value != "是") or \
               (key.startswith("不规范对端") and isinstance(value, int) and value > 0)
        _style_body_row(ws, i, 2, _WARN_FILL if warn else None)

    _autosize(ws, max_width=80)


def _write_hits_sheet(wb: Workbook, result: MultiGroupAnalysisResult):
    ws = wb.create_sheet("命中明细")

    headers = [
        "Destination",
        "路径数",
        "涉及组数",
        "涉及的平行设备组",
        "对端设备明细",
        "接口明细",
        "Pre",
        "Cost",
        "协议",
        "包含不规范对端",
    ]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    _style_header(ws, 1, len(headers))

    for i, hit in enumerate(result.hits, start=2):
        ws.cell(row=i, column=1, value=hit.destination)
        ws.cell(row=i, column=2, value=hit.path_count)
        ws.cell(row=i, column=3, value=hit.group_count)
        ws.cell(row=i, column=4, value="\n".join(hit.group_keys))
        ws.cell(row=i, column=5, value="\n".join(f"{p.group_key} / {p.peer_device}" for p in hit.paths))
        ws.cell(row=i, column=6, value="\n".join(f"{p.peer_device} ← {p.interface}" for p in hit.paths))
        ws.cell(row=i, column=7, value=",".join(sorted({str(p.pre) for p in hit.paths})))
        ws.cell(row=i, column=8, value=",".join(sorted({str(p.cost) for p in hit.paths})))
        ws.cell(row=i, column=9, value=",".join(hit.protocols))
        ws.cell(row=i, column=10, value="是" if hit.has_unparseable_peer else "")

        fill = _WARN_FILL if hit.has_unparseable_peer else _HIT_FILL
        _style_body_row(ws, i, len(headers), fill)

    ws.freeze_panes = "A2"
    _autosize(ws, max_width=70)


def _write_groups_sheet(wb: Workbook, result: MultiGroupAnalysisResult):
    ws = wb.create_sheet("平行设备组清单")

    headers = ["分组键", "成员数", "成员设备", "规范"]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    _style_header(ws, 1, len(headers))

    # 按"组成员数降序、再按分组键升序"
    sorted_groups = sorted(
        result.group_members.items(),
        key=lambda kv: (-len(kv[1]), kv[0]),
    )

    for i, (group_key, members) in enumerate(sorted_groups, start=2):
        ws.cell(row=i, column=1, value=group_key)
        ws.cell(row=i, column=2, value=len(members))
        ws.cell(row=i, column=3, value="\n".join(sorted(members)))
        is_unparseable_only = all(m in result.unparseable_peers for m in members)
        ws.cell(row=i, column=4, value="否" if is_unparseable_only else "是")
        _style_body_row(ws, i, len(headers), _WARN_FILL if is_unparseable_only else None)

    ws.freeze_panes = "A2"
    _autosize(ws, max_width=60)
