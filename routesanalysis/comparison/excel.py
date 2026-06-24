"""
Excel报告导出器 - 生成BGP路由比较报告
"""

import os
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import sys

# 尝试导入openpyxl，如果失败则提供友好的错误信息
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.error(f"openpyxl导入失败: {e}. 请安装: pip install openpyxl")

from routesanalysis.models import ComparisonResult, RouteDifference, DifferenceType

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Excel报告导出器
    生成BGP路由比较报告，支持百万级差异的优化导出
    """

    # 颜色定义
    HEADER_BG_COLOR = "366092"  # 深蓝色
    HEADER_FONT_COLOR = "FFFFFF"  # 白色
    HIGHLIGHT_COLOR = "FFC7CE"  # 浅红色（用于高亮差异）
    SUCCESS_COLOR = "C6EFCE"  # 浅绿色（用于成功/匹配）
    INFO_COLOR = "F2F2F2"  # 浅灰色（用于信息行）

    # 列宽定义
    COLUMN_WIDTHS = {
        "A": 25,  # Destination
        "B": 15,  # 设备1
        "C": 15,  # 设备2
        "D": 20,  # 差异类型
        "E": 40,  # 详细信息
    }

    def __init__(self, write_only: bool = True, batch_size: int = 10000):
        """
        初始化导出器

        Args:
            write_only: 是否使用write-only模式（提高大文件性能）
            batch_size: 分批写入的批量大小
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl未安装，无法导出Excel报告。请安装: pip install openpyxl")

        self.write_only = write_only
        self.batch_size = batch_size
        self._styles_initialized = False
        self._styles: Dict[str, Any] = {}

    def export(self, result: ComparisonResult, output_path: str,
               include_summary: bool = True, include_details: bool = True):
        """
        导出比较结果到Excel

        Args:
            result: 比较结果对象
            output_path: 输出文件路径
            include_summary: 是否包含汇总信息
            include_details: 是否包含详细差异

        Raises:
            ValueError: 输出路径无效或导出失败
        """
        logger.info(f"开始导出Excel报告到: {output_path}")

        # 验证输出路径
        output_path = self._normalize_output_path(output_path)

        try:
            if self.write_only and len(result.differences) > 10000:
                # 对于大量差异，使用write-only模式
                self._export_write_only(result, output_path, include_summary, include_details)
            else:
                # 对于少量差异，使用常规模式（支持格式）
                self._export_regular(result, output_path, include_summary, include_details)

            logger.info(f"Excel报告导出完成: {output_path}")

        except Exception as e:
            logger.error(f"导出Excel报告失败: {e}")
            raise ValueError(f"导出失败: {e}")

    def _export_regular(self, result: ComparisonResult, output_path: str,
                        include_summary: bool, include_details: bool):
        """常规模式导出（支持完整格式）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "BGP路由比较报告"

        # 初始化样式
        self._initialize_styles()

        current_row = 1

        # 写入报告标题
        current_row = self._write_report_title(ws, current_row, result)

        if include_summary:
            # 写入汇总信息
            current_row = self._write_summary_section(ws, current_row, result)

        if include_details and result.differences:
            # 写入详细差异
            current_row = self._write_details_section(ws, current_row, result)

        # 调整列宽
        self._adjust_column_widths(ws)

        # 保存文件
        wb.save(output_path)

    def _export_write_only(self, result: ComparisonResult, output_path: str,
                           include_summary: bool, include_details: bool):
        """Write-only模式导出（高性能，适用于大量数据）"""
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="BGP路由比较报告")

        # 在write-only模式下，不能使用样式，所以使用简单格式
        current_row = 1

        # 写入报告标题（简单文本）
        ws.append(["BGP路由比较报告", "", "", "", ""])
        ws.append([f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", "", "", "", ""])
        ws.append([f"基准设备: {result.summary['baseline_device']}", "", "", "", ""])
        ws.append([])
        current_row += 4

        if include_summary:
            # 写入汇总信息（简单文本）
            summary_rows = self._format_summary_text(result)
            for row in summary_rows:
                ws.append(row)
                current_row += 1
            ws.append([])  # 空行
            current_row += 1

        if include_details and result.differences:
            # 写入表头
            header = ["Destination", "设备1", "设备2", "差异类型", "详细信息"]
            ws.append(header)

            # 分批写入差异数据
            differences_batches = self._batch_differences(result.differences)
            for batch in differences_batches:
                for diff in batch:
                    row = self._format_difference_row(diff)
                    ws.append(row)

        # 保存文件
        wb.save(output_path)

    def _write_report_title(self, ws, start_row: int, result: ComparisonResult) -> int:
        """写入报告标题"""
        row = start_row

        # 主标题
        ws.cell(row=row, column=1, value="BGP路由比较报告")
        title_cell = ws.cell(row=row, column=1)
        title_cell.font = self._styles["title_font"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

        # 生成时间
        ws.cell(row=row, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        time_cell = ws.cell(row=row, column=1)
        time_cell.font = self._styles["info_font"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

        # 基准设备
        ws.cell(row=row, column=1, value=f"基准设备: {result.summary['baseline_device']}")
        baseline_cell = ws.cell(row=row, column=1)
        baseline_cell.font = self._styles["info_font"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 2  # 空一行

        return row

    def _write_summary_section(self, ws, start_row: int, result: ComparisonResult) -> int:
        """写入汇总信息部分"""
        row = start_row

        # 汇总标题
        ws.cell(row=row, column=1, value="=== 比较汇总 ===")
        summary_title_cell = ws.cell(row=row, column=1)
        summary_title_cell.font = self._styles["section_title_font"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

        # 基本信息
        ws.cell(row=row, column=1, value="基准设备:")
        ws.cell(row=row, column=2, value=result.summary['baseline_device'])
        row += 1

        ws.cell(row=row, column=1, value="比较设备:")
        compared_devices = ", ".join(result.summary['compared_devices'])
        ws.cell(row=row, column=2, value=compared_devices)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1

        ws.cell(row=row, column=1, value="基准设备路由总数:")
        ws.cell(row=row, column=2, value=result.summary['total_routes_baseline'])
        row += 1

        ws.cell(row=row, column=1, value="差异总数:")
        ws.cell(row=row, column=2, value=result.summary['total_differences'])
        row += 2  # 空一行

        # 差异类型统计
        ws.cell(row=row, column=1, value="差异类型统计:")
        stat_title_cell = ws.cell(row=row, column=1)
        stat_title_cell.font = self._styles["bold_font"]
        row += 1

        # 表头
        headers = ["差异类型", "数量"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self._styles["header_font"]
            cell.fill = self._styles["header_fill"]
            cell.alignment = self._styles["center_align"]
            cell.border = self._styles["thin_border"]

        row += 1

        # 差异类型数据
        diff_type_names = {
            "missing_destination": "缺少Destination",
            "missing_interface": "接口缺失",
            "interface_mismatch": "接口不同",
            "pre_cost_diff": "Pre/Cost差异"
        }

        for diff_type_key, count in result.summary['differences_by_type'].items():
            type_name = diff_type_names.get(diff_type_key, diff_type_key)

            ws.cell(row=row, column=1, value=type_name)
            ws.cell(row=row, column=2, value=count)

            # 添加边框
            for col in [1, 2]:
                cell = ws.cell(row=row, column=col)
                cell.border = self._styles["thin_border"]

            row += 1

        row += 2  # 空两行

        return row

    def _write_details_section(self, ws, start_row: int, result: ComparisonResult) -> int:
        """写入详细差异部分"""
        row = start_row

        # 详细差异标题
        ws.cell(row=row, column=1, value="=== 详细差异列表 ===")
        details_title_cell = ws.cell(row=row, column=1)
        details_title_cell.font = self._styles["section_title_font"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

        # 表头
        headers = ["Destination", "基准设备", "对比设备", "差异类型", "详细信息"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self._styles["header_font"]
            cell.fill = self._styles["header_fill"]
            cell.alignment = self._styles["center_align"]
            cell.border = self._styles["thin_border"]

        row += 1

        # 差异类型名称映射
        diff_type_names = {
            DifferenceType.MISSING_DESTINATION: "缺少Destination",
            DifferenceType.MISSING_INTERFACE: "接口缺失",
            DifferenceType.INTERFACE_MISMATCH: "接口不同",
            DifferenceType.PRE_COST_DIFFERENCE: "Pre/Cost差异"
        }

        # 分批写入差异数据（避免内存问题）
        differences_batches = self._batch_differences(result.differences)
        total_batches = len(differences_batches)

        for batch_idx, batch in enumerate(differences_batches, 1):
            for diff in batch:
                # 写入差异行
                ws.cell(row=row, column=1, value=diff.destination)
                ws.cell(row=row, column=2, value=diff.device1)
                ws.cell(row=row, column=3, value=diff.device2)

                # 差异类型（使用中文名称）
                type_name = diff_type_names.get(diff.difference_type, diff.difference_type.value)
                ws.cell(row=row, column=4, value=type_name)

                # 详细信息（格式化）
                details_text = self._format_difference_details(diff)
                ws.cell(row=row, column=5, value=details_text)

                # 根据差异类型设置背景色
                if diff.difference_type == DifferenceType.MISSING_DESTINATION:
                    fill_color = self.HIGHLIGHT_COLOR
                elif diff.difference_type == DifferenceType.MISSING_INTERFACE:
                    fill_color = "FFF2CC"  # 浅黄色
                elif diff.difference_type == DifferenceType.INTERFACE_MISMATCH:
                    fill_color = "FFE0B2"  # 淡橙色
                else:  # PRE_COST_DIFFERENCE
                    fill_color = "E2EFDA"  # 浅绿色

                # 应用背景色
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.border = self._styles["thin_border"]

                row += 1

            # 显示进度（每批）
            if total_batches > 1:
                logger.debug(f"写入差异数据: 批次 {batch_idx}/{total_batches}")

        return row

    def _format_summary_text(self, result: ComparisonResult) -> List[List[str]]:
        """格式化汇总信息为文本行（用于write-only模式）"""
        rows = []

        # 汇总标题
        rows.append(["=== 比较汇总 ===", "", "", "", ""])

        # 基本信息
        rows.append(["基准设备:", result.summary['baseline_device'], "", "", ""])
        rows.append(["比较设备:", ", ".join(result.summary['compared_devices']), "", "", ""])
        rows.append(["基准设备路由总数:", str(result.summary['total_routes_baseline']), "", "", ""])
        rows.append(["差异总数:", str(result.summary['total_differences']), "", "", ""])
        rows.append(["", "", "", "", ""])  # 空行

        # 差异类型统计
        rows.append(["差异类型统计:", "", "", "", ""])
        rows.append(["差异类型", "数量", "", "", ""])

        diff_type_names = {
            "missing_destination": "缺少Destination",
            "missing_interface": "接口缺失",
            "interface_mismatch": "接口不同",
            "pre_cost_diff": "Pre/Cost差异"
        }

        for diff_type_key, count in result.summary['differences_by_type'].items():
            type_name = diff_type_names.get(diff_type_key, diff_type_key)
            rows.append([type_name, str(count), "", "", ""])

        rows.append(["", "", "", "", ""])  # 空行

        return rows

    def _format_difference_row(self, diff: RouteDifference) -> List[str]:
        """格式化差异行（用于write-only模式）"""
        # 差异类型名称映射
        diff_type_names = {
            DifferenceType.MISSING_DESTINATION: "缺少Destination",
            DifferenceType.MISSING_INTERFACE: "接口缺失",
            DifferenceType.INTERFACE_MISMATCH: "接口不同",
            DifferenceType.PRE_COST_DIFFERENCE: "Pre/Cost差异"
        }

        type_name = diff_type_names.get(diff.difference_type, diff.difference_type.value)
        details_text = self._format_difference_details(diff)

        return [
            diff.destination,
            diff.device1,
            diff.device2,
            type_name,
            details_text
        ]

    def _format_difference_details(self, diff: RouteDifference) -> str:
        """格式化差异详细信息"""
        if diff.difference_type == DifferenceType.MISSING_DESTINATION:
            return f"在{diff.details['missing_in']}中缺少此Destination"
        elif diff.difference_type == DifferenceType.MISSING_INTERFACE:
            label = "对端设备" if diff.details.get("compared_by") == "peer_device" else "接口"
            return f"{label} {diff.details['interface']} 在 {diff.details['missing_in']} 中缺失"
        elif diff.difference_type == DifferenceType.INTERFACE_MISMATCH:
            details = diff.details
            return (f"接口列表不一致: "
                   f"{diff.device1}[{', '.join(details['device1_interfaces'])}] vs "
                   f"{diff.device2}[{', '.join(details['device2_interfaces'])}]")
        elif diff.difference_type == DifferenceType.PRE_COST_DIFFERENCE:
            details = diff.details
            return (f"接口 {details['interface']}: "
                   f"Pre({details['device1_pre']} vs {details['device2_pre']}), "
                   f"Cost({details['device1_cost']} vs {details['device2_cost']})")
        else:
            return str(diff.details)

    def _batch_differences(self, differences: List[RouteDifference]) -> List[List[RouteDifference]]:
        """将差异列表分批（用于批量写入）"""
        if not differences:
            return []

        batches = []
        for i in range(0, len(differences), self.batch_size):
            batch = differences[i:i + self.batch_size]
            batches.append(batch)

        logger.debug(f"差异数据分批: {len(differences)} 条差异分为 {len(batches)} 批")
        return batches

    def _initialize_styles(self):
        """初始化样式定义"""
        if self._styles_initialized:
            return

        # 字体定义
        self._styles["title_font"] = Font(name="微软雅黑", size=16, bold=True)
        self._styles["section_title_font"] = Font(name="微软雅黑", size=12, bold=True)
        self._styles["header_font"] = Font(name="微软雅黑", size=10, bold=True, color=self.HEADER_FONT_COLOR)
        self._styles["bold_font"] = Font(name="微软雅黑", size=10, bold=True)
        self._styles["info_font"] = Font(name="微软雅黑", size=10)

        # 填充定义
        self._styles["header_fill"] = PatternFill(
            start_color=self.HEADER_BG_COLOR,
            end_color=self.HEADER_BG_COLOR,
            fill_type="solid"
        )

        # 对齐定义
        self._styles["center_align"] = Alignment(horizontal="center", vertical="center")
        self._styles["left_align"] = Alignment(horizontal="left", vertical="center")

        # 边框定义
        thin_border = Side(border_style="thin", color="000000")
        self._styles["thin_border"] = Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border
        )

        self._styles_initialized = True

    def _adjust_column_widths(self, ws):
        """调整列宽"""
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        # 自动调整其他列（使用列索引避免合并单元格问题）
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter not in self.COLUMN_WIDTHS:
                max_length = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (AttributeError, TypeError):
                        continue
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

    def _normalize_output_path(self, output_path: str) -> str:
        """标准化输出路径"""
        path = Path(output_path)

        # 确保目录存在
        path.parent.mkdir(parents=True, exist_ok=True)

        # 如果没有扩展名，添加.xlsx
        if not path.suffix:
            path = path.with_suffix('.xlsx')

        # Windows长路径处理
        if sys.platform == "win32":
            if len(str(path)) > 260:
                long_path = "\\\\?\\" + str(path)
                return long_path

        return str(path)


# 快捷函数
def export_comparison_result(result: ComparisonResult, output_path: str,
                            write_only: bool = True) -> str:
    """
    导出比较结果到Excel（快捷函数）

    Args:
        result: 比较结果对象
        output_path: 输出文件路径
        write_only: 是否使用write-only模式

    Returns:
        输出文件路径
    """
    exporter = ExcelExporter(write_only=write_only)
    exporter.export(result, output_path)
    return output_path


def export_differences_to_excel(differences: List[RouteDifference], output_path: str,
                               baseline_name: str = "设备1",
                               compared_name: str = "设备2") -> str:
    """
    将差异列表导出到Excel（快捷函数）

    Args:
        differences: 差异列表
        output_path: 输出文件路径
        baseline_name: 基准设备名称
        compared_name: 比较设备名称

    Returns:
        输出文件路径
    """
    from routesanalysis.models import Device, ComparisonResult

    # 创建虚拟设备（用于满足ComparisonResult接口）
    dummy_device = Device(name=baseline_name, filename="", routes=[])
    dummy_compared = [Device(name=compared_name, filename="", routes=[])]

    # 创建虚拟汇总信息
    summary = {
        "baseline_device": baseline_name,
        "compared_devices": [compared_name],
        "total_routes_baseline": 0,
        "total_differences": len(differences),
        "differences_by_type": {},
        "differences_by_device_pair": {},
        "performance_stats": {}
    }

    # 统计差异类型
    from collections import Counter
    diff_counter = Counter()
    for diff in differences:
        diff_counter[diff.difference_type] += 1
    summary["differences_by_type"] = {k.value: v for k, v in diff_counter.items()}

    result = ComparisonResult(
        baseline_device=dummy_device,
        compared_devices=dummy_compared,
        differences=differences,
        summary=summary
    )

    # 导出
    return export_comparison_result(result, output_path)